"""
Grid report engine
===================
Regenerates a dashboard grid server-side (no browser) so it can be emailed on a
schedule. A "grid" scheduled report stores, at creation time:

    { "kind": "grid",
      "endpoint": "/api/purchases/details",     # the same API the grid uses
      "params":   { "stores": "...", "vendors": "..." },   # slicer values
      "period":   "MTD",                         # 30D|90D|7D|MTD|YTD|custom
      "date_from": "2026-07-01", "date_to": "2026-07-15",  # used when period=custom
      "columns":  [ { "id": "vou_date", "label": "Date" }, ... ],
      "title": "...", "view": "...", "filters": "...", "name": "..." }

At send time `run_grid()` calls the endpoint's Python function directly (the
scheduler already runs in-process with DuckDB access, so no HTTP / auth round
trip is needed — supplying the kwargs explicitly bypasses FastAPI's Depends()).
The rows are turned into a CSV attachment plus an HTML preview table.

To make a new grid schedulable: add its endpoint path → function to REGISTRY.
"""
import csv
import io
import os
import logging
import inspect
from datetime import date, datetime, timedelta

try:
    from fastapi import params as _fparams
except Exception:                       # pragma: no cover
    _fparams = None

log = logging.getLogger(__name__)

# Hard cap so a scheduled email can never balloon. Raised from 20,000 on
# 19 Jul 2026: a Coverage export is one row per item x store (~74k rows here),
# so the old cap silently truncated the sheet. CSV/Excel handle this size fine;
# PDF stays capped separately at _PDF_MAX_ROWS. Override with RETAILTEC_MAX_REPORT_ROWS.
_MAX_ROWS = int(os.environ.get("RETAILTEC_MAX_REPORT_ROWS", "500000"))
_PREVIEW_ROWS = 30         # rows shown inline in the email body


def max_report_rows() -> int:
    """Effective cap: Settings → Reports & Email wins, else the default above.

    Admin-editable so a site can raise/lower it without a rebuild. Never let a
    bad stored value collapse the cap to something that truncates silently.
    """
    try:
        from services.config import load_settings
        v = int((load_settings().get("email") or {}).get("max_report_rows") or 0)
        if v > 0:
            return v
    except Exception:
        pass
    return _MAX_ROWS


# ── Endpoint registry (path → callable) ──────────────────────────────────────

def _registry() -> dict:
    """Lazy import to avoid circular imports at module load."""
    from routers import purchases, sales, inventory, accounting
    reg = {
        # Sales / purchasing
        "/api/purchases/details":       purchases.purchases_details,
        "/api/sales/transactions":      sales.transactions,
        "/api/sales/products":          sales.products,
        "/api/sales/journal/invoices":  getattr(sales, "journal_invoices", None),
        "/api/sales/journal/items":     getattr(sales, "journal_items", None),
        "/api/sales/perf/associates":   getattr(sales, "perf_associates", None),
        "/api/sales/perf/customers":    getattr(sales, "perf_customers", None),
        "/api/sales/perf/stores":       getattr(sales, "perf_stores", None),
        # Inventory
        "/api/inventory/items":         getattr(inventory, "inv_items", None),
        "/api/inventory/movement-by":   getattr(inventory, "inv_movement_by", None),
        "/api/inventory/transfers/details":   getattr(inventory, "transfers_details", None),
        "/api/inventory/adjustments/details": getattr(inventory, "adjustments_details", None),
        "/api/inventory/ledger":        getattr(inventory, "inventory_ledger", None),
        "/api/inventory/coverage":      getattr(inventory, "inv_coverage", None),
        "/api/inventory/history/details": getattr(inventory, "invh_details", None),
        "/api/inventory/stock-asof":    getattr(inventory, "stock_asof", None),
        # Accounting (virtual GL). /api/accounting/journal is not listed: it
        # returns {total, rows}, not a bare list, so it is not a grid the report
        # engine can render — the lines grid covers the same data as rows.
        "/api/accounting/journal/lines": getattr(accounting, "gl_journal_lines", None),
        "/api/accounting/trial-balance": getattr(accounting, "gl_trial_balance", None),
        "/api/accounting/profit-loss":  getattr(accounting, "gl_profit_loss", None),
        "/api/accounting/balance-sheet": getattr(accounting, "gl_balance_sheet", None),
        "/api/accounting/general-ledger": getattr(accounting, "gl_general_ledger", None),
        # bp-statement replays with the partner SID riding params["bp_id"];
        # aging replays with params["as_of"] / ["side"] / ["buckets"].
        "/api/accounting/bp-statement": getattr(accounting, "gl_bp_statement", None),
        "/api/accounting/aging":        getattr(accounting, "gl_aging", None),
        "/api/accounting/exceptions":   getattr(accounting, "gl_exceptions", None),
    }
    return {k: v for k, v in reg.items() if v is not None}


def endpoint_supported(endpoint: str) -> bool:
    try:
        return endpoint in _registry()
    except Exception:
        return False


# ── Period → date window ──────────────────────────────────────────────────────

def resolve_window(report: dict) -> tuple[str, str]:
    """Return (date_from, date_to) ISO strings. A preset rolls relative to today
    so a recurring report always covers the intended trailing window; 'custom'
    (or unknown) uses the stored absolute dates."""
    today = date.today()
    period = (report.get("period") or "").upper()

    def iso(d): return d.isoformat()

    if period == "30D":
        return iso(today - timedelta(days=29)), iso(today)
    if period == "90D":
        return iso(today - timedelta(days=89)), iso(today)
    if period == "7D":
        return iso(today - timedelta(days=6)), iso(today)
    if period == "MTD":
        return iso(today.replace(day=1)), iso(today)
    if period == "YTD":
        return iso(today.replace(month=1, day=1)), iso(today)
    # custom / unknown → stored absolute dates (fall back to last 30 days)
    df = report.get("date_from") or iso(today - timedelta(days=29))
    dt = report.get("date_to") or iso(today)
    return df, dt


# ── Run the grid's data function ──────────────────────────────────────────────

def _coerce(name: str, value):
    """date_from / date_to / as_of arrive as ISO strings; endpoint signatures
    type them as datetime.date, so convert. (as_of: the Balance Sheet's single
    cut-off date — scheduled BS reports replay the stored absolute date.)"""
    if name in ("date_from", "date_to", "as_of") and isinstance(value, str):
        try:
            return date.fromisoformat(value[:10])
        except Exception:
            return value
    return value


def _safe_default(param):
    """Resolve a parameter's default to a *real* value for a direct (non-HTTP)
    call. FastAPI endpoints declare defaults like Depends(scoped_stores) or
    Query(None); those marker objects must never reach the function body (e.g.
    store_filter would call .split() on a Depends object). Depends → None
    (no scoping = all stores); Query/Path/etc → their underlying default."""
    d = param.default
    if _fparams is not None:
        if isinstance(d, _fparams.Depends):
            return None
        if isinstance(d, _fparams.Param):          # Query / Path / Header / Cookie
            val = getattr(d, "default", None)
            if val is inspect._empty or val is ... or type(val).__name__ == "PydanticUndefinedType":
                return None
            return val
    if d is inspect._empty:
        return None
    return d


def run_grid(endpoint: str, params: dict) -> list[dict]:
    reg = _registry()
    fn = reg.get(endpoint)
    if fn is None:
        raise RuntimeError(f"Endpoint not schedulable: {endpoint}")
    sig = inspect.signature(fn)
    supplied = params or {}
    kwargs = {}
    for name, param in sig.parameters.items():
        if param.kind in (inspect.Parameter.VAR_POSITIONAL,
                           inspect.Parameter.VAR_KEYWORD):
            continue
        if name in supplied and supplied[name] is not None:
            kwargs[name] = _coerce(name, supplied[name])
        else:
            # Not provided (or provided as None) → use a safe concrete default,
            # never a leftover FastAPI Depends()/Query() marker object.
            kwargs[name] = _safe_default(param)
    rows = fn(**kwargs)
    # Endpoint functions return list[dict] (qdf) or occasionally a dict; normalise
    if isinstance(rows, dict):
        rows = rows.get("rows") or rows.get("data") or []
    rows = list(rows or [])
    cap = max_report_rows()
    if len(rows) > cap:
        log.warning("Report %s truncated: %d rows -> cap %d", endpoint, len(rows), cap)
        rows = rows[:cap]
    return rows


# ── Item identifier (server-side mirror of Settings → Display) ───────────────
# The frontend stores which DIM_ITEM field identifies an item (ALU / UPC /
# Description) and resolves blanks to ALU with a grid valueGetter. Attachments
# are built server-side, so the same rules live here, driven by
# settings.json → display.item_identifier (synced by the frontend).

_IDENT_FID = "__item_id"        # synthetic field for the resolved identifier
_IDENT_FAMILY = {"alu", "upc", "description", "description1"}
_IDENT_LABELS = {"alu": "ALU", "upc": "UPC", "description": "Description"}


def item_identifier_cfg() -> tuple[str, str]:
    """Configured identifier → (field, label). Whitelisted; default ALU."""
    try:
        from services.config import load_settings
        v = str(((load_settings().get("display") or {}).get("item_identifier"))
                or "alu").lower()
    except Exception:
        v = "alu"
    if v not in _IDENT_LABELS:
        v = "alu"
    return v, _IDENT_LABELS[v]


def item_identifier_sql(alias: str = "I") -> tuple[str, str]:
    """(SQL expression, label) for the configured identifier on a DIM_ITEM
    alias, falling back to ALU when the field is blank. Column names come from
    a fixed internal whitelist, never from user input."""
    field, label = item_identifier_cfg()
    col = {"alu": "ALU", "upc": "UPC", "description": "DESCRIPTION1"}[field]
    if col == "ALU":
        return f"{alias}.ALU", label
    return f"COALESCE(NULLIF({alias}.{col}, ''), {alias}.ALU)", label


def _row_key(rows: list[dict], name: str):
    """Actual row key for a logical field name — both casings exist across
    endpoints, and 'description' is aliased 'description1' on some grids."""
    if not rows:
        return None
    keys = rows[0].keys()
    cands = [name, name.upper()]
    if name == "description":
        cands += ["description1", "DESCRIPTION1"]
    for c in cands:
        if c in keys:
            return c
    return None


def _apply_item_identifier(cols: list[tuple[str, str]],
                           rows: list[dict]) -> list[tuple[str, str]]:
    """Collapse the alu/upc/description column family into ONE identifier
    column: the configured field with fallback to ALU when blank (same as the
    frontend valueGetter), headed with the configured label. A Description
    column that is NOT the identifier stays — it is distinct data the report
    showed; the identifier itself never appears twice."""
    if not any(fid.lower() in _IDENT_FAMILY for fid, _ in cols):
        return cols
    field, label = item_identifier_cfg()
    id_key = _row_key(rows, field)
    alu_key = _row_key(rows, "alu")
    out, inserted = [], False
    for fid, lbl in cols:
        low = fid.lower()
        if low not in _IDENT_FAMILY:
            out.append((fid, lbl))
            continue
        if not inserted:
            out.append((_IDENT_FID, label))
            inserted = True
        if low in ("description", "description1") and field != "description":
            out.append((fid, lbl))       # keep Description as its own data
    for r in rows:
        v = r.get(id_key) if id_key else None
        if v is None or str(v).strip() == "":
            v = r.get(alu_key) if alu_key else v
        r[_IDENT_FID] = "" if v is None else v
    return out


# ── Output builders ───────────────────────────────────────────────────────────

def _columns(report: dict, rows: list[dict]) -> list[tuple[str, str]]:
    """Return [(field, label)]. Uses stored columns; falls back to row keys."""
    cols = report.get("columns") or []
    out = []
    for c in cols:
        fid = c.get("id") or c.get("field")
        if fid:
            out.append((fid, c.get("label") or fid))
    if not out and rows:
        out = [(k, k) for k in rows[0].keys() if k != _IDENT_FID]
    # Drop presentation-only columns with no backing data — e.g. the pinned '#'
    # row-number column some grids show (its colId matches no row key), which
    # rendered as an empty leading column in every attachment format.
    if rows:
        keys = set(rows[0].keys())
        out = [(fid, lbl) for fid, lbl in out if fid in keys or fid == _IDENT_FID]
    # ONE item identifier column per Settings → Display, with ALU fallback.
    out = _apply_item_identifier(out, rows)
    return out


def build_csv(report: dict, rows: list[dict]) -> bytes:
    cols = _columns(report, rows)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([label for _, label in cols])
    for r in rows:
        w.writerow([r.get(fid, "") for fid, _ in cols])
    return buf.getvalue().encode("utf-8-sig")   # BOM → Excel opens UTF-8 cleanly


def build_xlsx(report: dict, rows: list[dict]) -> bytes:
    """Build an .xlsx workbook (bold header, frozen top row, auto-ish widths)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    cols = _columns(report, rows)
    wb = Workbook()
    ws = wb.active
    ws.title = (report.get("title") or report.get("name") or "Report")[:31]
    hdr_fill = PatternFill("solid", fgColor="EEF2FF")
    hdr_font = Font(bold=True, color="3730A3")
    for ci, (_, label) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=ci, value=label)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(vertical="center")
    for ri, r in enumerate(rows, start=2):
        for ci, (fid, _) in enumerate(cols, start=1):
            v = r.get(fid)
            if isinstance(v, (list, dict)):
                v = str(v)
            ws.cell(row=ri, column=ci, value=v)
    # column widths from header + a sample of the data
    for ci, (fid, label) in enumerate(cols, start=1):
        w = len(str(label))
        for r in rows[:200]:
            w = max(w, len(str(r.get(fid, ""))))
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max(w + 2, 8), 48)
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_PDF_MAX_ROWS = 1500       # keep a scheduled PDF a sane size


def _pdf_font(pdf):
    """Register a Unicode TTF so Arabic/accented text doesn't crash the core
    latin-1 fonts. Falls back to Helvetica (latin-1) if no TTF is found."""
    import os
    for path in (r"C:\Windows\Fonts\arial.ttf", r"C:\Windows\Fonts\segoeui.ttf",
                 r"C:\Windows\Fonts\tahoma.ttf"):
        if os.path.exists(path):
            try:
                pdf.add_font("uni", "", path)
                pdf.add_font("uni", "B", path)
                return "uni", True
            except Exception:
                pass
    return "helvetica", False


def build_pdf(report: dict, rows: list[dict]) -> bytes:
    """Landscape A4 table PDF of the grid (header repeats on each page)."""
    from fpdf import FPDF
    cols = _columns(report, rows)
    title = report.get("title") or report.get("name") or "Report"

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(True, margin=10)
    pdf.add_page()
    font, unicode_ok = _pdf_font(pdf)

    def txt(v):
        s = "" if v is None else str(v)
        return s if unicode_ok else s.encode("latin-1", "replace").decode("latin-1")

    pdf.set_font(font, "B", 13)
    pdf.cell(0, 7, txt(title), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font(font, "", 8)
    pdf.set_text_color(120, 120, 120)
    sub = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}  ·  {len(rows):,} rows"
    if len(rows) > _PDF_MAX_ROWS:
        sub += f"  (first {_PDF_MAX_ROWS:,} shown — full data in CSV/Excel)"
    pdf.cell(0, 5, txt(sub), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(1)
    pdf.set_text_color(15, 23, 42)

    data = [[txt(c[1]) for c in cols]] + \
           [[txt(r.get(fid)) for fid, _ in cols] for r in rows[:_PDF_MAX_ROWS]]
    pdf.set_font(font, "", 6.5)
    with pdf.table(first_row_as_headings=True, line_height=4.2,
                   text_align="LEFT", padding=(0.6, 1.2)) as table:
        for i, drow in enumerate(data):
            trow = table.row()
            for cell in drow:
                trow.cell(cell)
    out = pdf.output()
    return bytes(out)


def build_attachment(report: dict, rows: list[dict], fmt: str):
    """Return (bytes, mime, extension) for the requested attachment format."""
    fmt = (fmt or "csv").lower()
    if fmt in ("xlsx", "excel"):
        return build_xlsx(report, rows), \
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"
    if fmt == "pdf":
        return build_pdf(report, rows), "application/pdf", "pdf"
    return build_csv(report, rows), "text/csv", "csv"


def build_preview_html(report: dict, rows: list[dict]) -> str:
    cols = _columns(report, rows)
    if not cols:
        return "<p style='color:#94a3b8'>No rows for the selected period.</p>"
    head = "".join(
        f"<th style='text-align:left;padding:6px 10px;background:#f1f5f9;"
        f"border-bottom:2px solid #e2e8f0;font-size:12px;color:#475569'>{label}</th>"
        for _, label in cols)
    body = ""
    for r in rows[:_PREVIEW_ROWS]:
        tds = "".join(
            f"<td style='padding:5px 10px;border-bottom:1px solid #eef2f7;"
            f"font-size:12px;color:#0f172a'>{'' if r.get(fid) is None else r.get(fid)}</td>"
            for fid, _ in cols)
        body += f"<tr>{tds}</tr>"
    more = ""
    if len(rows) > _PREVIEW_ROWS:
        more = (f"<p style='margin:8px 0 0;color:#64748b;font-size:12px'>"
                f"Showing first {_PREVIEW_ROWS} of {len(rows):,} rows — full data in the attached CSV.</p>")
    return (f"<div style='overflow-x:auto'><table style='width:100%;border-collapse:collapse'>"
            f"<thead><tr>{head}</tr></thead><tbody>{body}</tbody></table></div>{more}")
